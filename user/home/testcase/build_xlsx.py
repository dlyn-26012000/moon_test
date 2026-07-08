#!/usr/bin/env python3
"""Sinh home-testcases.xlsx (core suite, 45 case). Chạy: python3 build_xlsx.py"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "home-testcases.xlsx")

# (ID, Tiêu đề, Loại, Dữ liệu/Bước, Kỳ vọng, Kết quả)
ROWS = [
    ("TC-F01","Mở Home (guest) render đủ section","Functional-Guest","GET /","6 section hiển thị (trừ section ẩn khi rỗng)","⬜"),
    ("TC-F02","Hero 3 carousel autoplay & loop","Functional-Guest","quan sát","main 3s/side 4s/side 5s, loop","⬜"),
    ("TC-F03","Click banner -> banner.link","Functional-Guest","click ảnh","tới đúng URL","⬜"),
    ("TC-F04","FlashSale countdown chạy lùi","Functional-Guest","quan sát","h:m:s giảm dần","⬜"),
    ("TC-F05","FlashSale <=4 card + Xem tất cả -> /products?sale=1","Functional-Guest","xem grid + click","<=4 card, điều hướng đúng","⬜"),
    ("TC-F06","MostLiked card + link /favorites","Functional-Guest","xem","card + link đúng","⬜"),
    ("TC-F07","Featured <=4 + link /products?featured=1","Functional-Guest","click","điều hướng đúng","⬜"),
    ("TC-F08","ProductCard -> chi tiết /products/{slug}","Functional-Guest","click card","tới chi tiết","⬜"),
    ("TC-F09","Header search theo keyword","Functional-Guest","nhập + Enter","tới kết quả với keyword","⬜"),
    ("TC-F10","CTA Đăng nhập/Đăng ký mở modal","Functional-Guest","click account","modal mở","⬜"),
    ("TC-F11","Language switcher vi<->en đổi text","Functional-Guest","đổi ngôn ngữ","text đổi","⬜"),
    ("TC-L01","Login -> Header logged","Functional-Logged","login -> /","avatar+dropdown","⬜"),
    ("TC-L02","Nút tim phản ánh is_favorited","Functional-Logged","xem card đã like","tim active","⬜"),
    ("TC-L03","Toggle favorite từ Home","Functional-Logged","click tim","POST toggle 200, đổi trạng thái+count","⬜"),
    ("TC-L04","Guest->logged chỉ khác is_favorited","Functional-Logged","so sánh","nội dung section giống","⬜"),
    ("TC-L05","Logout -> guest","Functional-Logged","dropdown Logout","về CTA login, token xoá","⬜"),
    ("TC-A01","GET banners?includes=file","API","request","200, id/type/link/file.url, is_active, order asc","⬜"),
    ("TC-A02","GET products is_sale per_page=4","API","request","200 <=4 public, có prices/thumbnail","⬜"),
    ("TC-A03","GET products is_featured per_page=4","API","request","200 <=4 featured","⬜"),
    ("TC-A04","GET products/top-favorites?limit=8","API","request","200 sort favorites desc, count>0","⬜"),
    ("TC-A05","GET campaigns/active","API","request","200 campaign+products public+ends_at","⬜"),
    ("TC-A06","Envelope {success,code,data} + Content-Type","API","xem body/header","đúng shape & json","⬜"),
    ("TC-V01","top-favorites?limit=8 có tôn trọng?","Validation","limit=8","Nghi ngờ trả 10 (bỏ limit) BUG-C1","⬜"),
    ("TC-V02","products?per_page=0","Validation","0","422 min:1 hoặc default an toàn","⬜"),
    ("TC-V03","products?per_page=abc","Validation","abc","cast->0 422, không crash","⬜"),
    ("TC-V04","filters[is_sale][]=1 mảng","Validation","mảng","Nghi ngờ trim(array) 500 BUG-C2","⬜"),
    ("TC-V05","orders[password]=asc cột lạ","Validation","injection","bỏ qua fallback id desc","⬜"),
    ("TC-V06","orders[id]=;DROP TABLE","Validation","SQLi","sanitize desc, không thực thi","⬜"),
    ("TC-S01","5 endpoint public khi guest","Security","no token","200","⬜"),
    ("TC-S02","No IDOR Home","Security","field public","không lộ user khác","⬜"),
    ("TC-S03","ISR không rò rỉ is_favorited","Security","cache chung","không dữ liệu riêng (BUG-C5?)","⬜"),
    ("TC-S04","top-favorites không lộ PII","Security","chỉ product+count","đúng","⬜"),
    ("TC-S05","XSS qua banner.link/tên","Security","escape","không thực thi","⬜"),
    ("TC-S06","Security headers","Security","xem header","ghi nhận hiện trạng","⬜"),
    ("TC-S07","Method sai (POST vào GET)","Security","POST","405","⬜"),
    ("TC-E01","Không banner","Edge","empty","EmptyBanner, không lỗi","⬜"),
    ("TC-E02","Không sale & không campaign","Edge","empty","text no_flash_sale","⬜"),
    ("TC-E03","Không featured / không favorite","Edge","empty","Featured & MostLiked ẩn","⬜"),
    ("TC-E04","API 500/null 1 section","Edge","lỗi","section empty, phần khác render","⬜"),
    ("BUG-C1","top-favorites limit bị bỏ qua (default 10)","BugCandidate","limit=8 & limit=2, đếm item","limit không có tác dụng => bug","⬜"),
    ("BUG-C2","filters[is_x][] mảng -> trim(array) 500","BugCandidate","products?filters[is_sale][]=1","KHÔNG 500 là đạt","⬜"),
    ("BUG-C3","banner.link null -> Link lỗi","BugCandidate","banner link rỗng","không crash render","⬜"),
    ("BUG-C4","Campaign ends_at quá khứ vẫn active","BugCandidate","kiểm scope active + countdown","hành vi hết hạn hợp lý","⬜"),
    ("BUG-C5","ISR cache dùng chung lộ trạng thái user","BugCandidate","so sánh SSR guest vs logged","SSR không chứa is_favorited","⬜"),
    ("BUG-C6","SSR luôn language=vi","BugCandidate","tải Home locale en","HTML lần đầu vi","⬜"),
]

HDR = ["ID","Tiêu đề","Loại","Dữ liệu/Bước","Kỳ vọng","Kết quả"]

wb = Workbook(); ws = wb.active; ws.title = "Home Core Test Cases"
hf = Font(bold=True, color="FFFFFF"); hfill = PatternFill("solid", fgColor="1F4E78")
thin = Side(style="thin", color="CCCCCC"); border = Border(thin,thin,thin,thin)
wrap = Alignment(wrap_text=True, vertical="top")

ws.append(HDR)
for c in ws[1]:
    c.font = hf; c.fill = hfill; c.alignment = Alignment(horizontal="center", vertical="center"); c.border = border
for r in ROWS:
    ws.append(list(r))
for row in ws.iter_rows(min_row=2):
    for c in row:
        c.alignment = wrap; c.border = border
for i, w in enumerate([12, 44, 16, 34, 46, 10], 1):
    ws.column_dimensions[chr(64+i)].width = w
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:F{len(ROWS)+1}"
wb.save(OUT)
print(f"Wrote {OUT} with {len(ROWS)} cases")
