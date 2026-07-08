#!/usr/bin/env python3
"""Generate login-testcases.xlsx from the same case list as login-testcases.md."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

CASES = [
    # id, category, title, precondition, data, expected, result
    ("TC-F01","Functional","Login thành công (happy path)","user001 active","user001/password","200 LOGIN_SUCCESS + token","Pass"),
    ("TC-F02","Functional","Token dùng cho route bảo vệ","token F01","GET user/auth/me","200 profile","Pass"),
    ("TC-F03","Functional","Logout xoá token hiện tại","đã login","DELETE logout","200 LOGOUT_SUCCESS","Pass"),
    ("TC-F04","Functional","Multi-device 2 token","user001","login x2","2 token khác nhau hợp lệ","Pass"),
    ("TC-F05","Functional","Logout A không huỷ token B","2 token","logout A","B còn dùng được","Pass"),
    ("TC-F06","Functional","State transition guest->in->guest","-","login/logout","đúng trạng thái","Pass"),
    ("TC-F07","Functional","UI mở modal login","trang chủ","account->Đăng nhập","modal hiện, focus username","Pass"),
    ("TC-F08","Functional","UI login thành công","modal mở","nhập đúng submit","đóng modal, token lưu","Pass"),
    ("TC-F09","Functional","UI submit disabled khi thiếu input","modal mở","bỏ trống","nút disabled, không gọi API","Pass"),
    ("TC-F10","Functional","UI toggle hiện/ẩn mật khẩu","modal mở","click mắt","type text<->password","Pass"),
    ("TC-F11","Functional","UI link Quên mật khẩu","modal mở","click","tới /forgot-password","Pass"),
    ("TC-F12","Functional","UI chuyển sang Đăng ký","modal mở","click Đăng ký ngay","mở register modal","Pass"),
    ("TC-F13","Functional","UI đóng modal Esc/overlay/X","modal mở","Esc/click nền/X","đóng, khôi phục scroll","Pass"),
    ("TC-V01","Validation","Thiếu cả 2 field","-","{}","422 VALIDATION_ERROR","Pass"),
    ("TC-V02","Validation","Thiếu username","-","{password}","422 username required","Pass"),
    ("TC-V03","Validation","Thiếu password","-","{username}","422 password required","Pass"),
    ("TC-V04","Validation","Username rỗng","-","username=''","422 required","Pass"),
    ("TC-V05","Validation","Password rỗng","-","password=''","422 required","Pass"),
    ("TC-V06","Validation","Username toàn space (client)","modal","'   '","submit disabled, no API","Pass"),
    ("TC-V07","Validation","Space đầu/cuối username (API)","-","'  user001  '","200 SUCCESS - nay chuẩn hoa tuong minh (trim+LOWER)","Pass (BUG-006 fixed)"),
    ("TC-V08","Validation","Username là số","-","username=12345","Thực tế 422 must be a string (không coerce)","Pass"),
    ("TC-V09","Validation","Password là mảng","-","password=[...]","422 must be string","Pass"),
    ("TC-V10","Validation","Password null","-","password=null","422 required","Pass"),
    ("TC-V11","Validation","Unicode/emoji username","-","user001😀","USER_NOT_FOUND, no crash","Pass"),
    ("TC-V12","Validation","Chuỗi cực dài 5000","-","5000 'a'","no crash, USER_NOT_FOUND","Pass"),
    ("TC-V13","Validation","HTML/script username","-","<script>","an toàn, không thực thi","Pass"),
    ("TC-B01","Business","Đúng user + pass","active","user001/password","200 token","Pass"),
    ("TC-B02","Business","Đúng user + sai pass","-","user001/wrong","401 INVALID_CREDENTIALS","Pass (BUG-001 fixed)"),
    ("TC-B03","Business","User không tồn tại","-","nosuch","401 INVALID_CREDENTIALS (hết enumeration)","Pass (BUG-001/002 fixed)"),
    ("TC-B04","Business","User inactive","is_active=false","-","403 USER_INACTIVE","Blocked live (fix qua unit test)"),
    ("TC-B05","Business","User bị xoá cứng","no SoftDeletes","-","= USER_NOT_FOUND","N/A"),
    ("TC-B06","Business","Email chưa verify vẫn login","login không check","-","login vẫn thành công","Warning (OQ)"),
    ("TC-B07","Business","Remember me","không có tính năng","-","N/A token không hết hạn","N/A"),
    ("TC-B08","Business","Token expired","expiration null","-","không thể test","N/A"),
    ("TC-B09","Business","Login nhiều browser","-","-","mỗi phiên token riêng","Pass"),
    ("TC-B10","Business","Case-insensitive username","collation ci","USER001","có thể login","Pass (OQ-2)"),
    ("TC-B11","Business","Dùng lại token sau logout","-","-","401 UNAUTHENTICATED","Pass"),
    ("TC-S01","Security","SQLi username","user001' OR 1=1 --","-","bound param, no bypass","Pass"),
    ("TC-S02","Security","SQLi password","' OR '1'='1","-","Hash fail, no bypass","Pass"),
    ("TC-S03","Security","XSS qua message","<script>","-","React escape, no exec","Pass"),
    ("TC-S04","Security","Brute-force / Rate limit",">5/phút","-","429 Too Many Attempts","Pass"),
    ("TC-S05","Security","Rate-limit header","-","-","X-RateLimit-* đúng","Pass"),
    ("TC-S06","Security","User enumeration","-","so 2 message","2 message giống nhau -> hết lộ","Pass (BUG-002 fixed)"),
    ("TC-S07","Security","Token tampering","-","sửa token /me","401","Pass"),
    ("TC-S08","Security","Sai ability","-","token ability khác","403","Not-run"),
    ("TC-S09","Security","CSRF","Bearer stateless","-","N/A","N/A"),
    ("TC-S10","Security","Security headers","-","-","có XCTO/XFO/CSP/HSTS","Pass (BUG-005 fixed)"),
    ("TC-S11","Security","CORS wildcard","-","-","CORS siết ve *.dlyn.site","Pass (BUG-005 fixed)"),
    ("TC-S12","Security","Open redirect","không có param","-","N/A","N/A"),
    ("TC-S13","Security","Privilege escalation","-","user token -> admin","403/401","Not-run"),
    ("TC-S14","Security","Token trong localStorage","-","-","XSS đọc được","Warning (design)"),
    ("TC-S15","Security","HTTP status lỗi nghiệp vụ","-","auth-fail","401/403 đúng chuẩn","Pass (BUG-001 fixed)"),
    ("TC-P01","Performance","Response time login","-","-","< 1s (~0.5s)","Pass"),
    ("TC-P02","Performance","Concurrent login","-","-","rate-limit chặn tải","Warning"),
    ("TC-P03","Performance","Stress login","-","-","throttle ngăn stress 1 IP","Blocked"),
    ("TC-R01","Responsive","Desktop 1366x900","-","-","layout đúng","Pass"),
    ("TC-R02","Responsive","Tablet 768x1024","-","-","không tràn","Pass"),
    ("TC-R03","Responsive","Mobile 390x844","-","-","không overflow ngang","Pass"),
    ("TC-A01","Accessibility","Keyboard navigation","-","-","tab order đúng","Pass"),
    ("TC-A02","Accessibility","Focus quản lý","-","-","auto-focus username","Pass"),
    ("TC-A03","Accessibility","Tab order","-","-","nút mắt loại khỏi tab","Pass"),
    ("TC-A04","Accessibility","ARIA","-","-","role=dialog, aria-modal","Pass"),
    ("TC-A05","Accessibility","Đóng bằng Esc","-","-","Esc đóng modal","Pass"),
    ("TC-A06","Accessibility","Contrast","-","-","WCAG AA cần đo","Warning"),
    ("TC-A07","Accessibility","Screen reader error","-","-","có role=alert/aria-live","Pass (BUG-004 fixed)"),
    ("TC-L01","Multi-language","Nhãn VI","-","-","khớp vi/auth.json","Pass"),
    ("TC-L02","Multi-language","Nhãn EN","-","-","khớp en/auth.json","Pass"),
    ("TC-L03","Multi-language","Header language gửi API","-","-","đúng ngôn ngữ","Pass"),
    ("TC-L04","Multi-language","Message lỗi chưa i18n","-","-","đã map câu dịch","Pass (BUG-003 fixed)"),
    ("TC-L05","Multi-language","Tràn chữ VI/EN","-","-","không tràn","Pass"),
]

HEAD = ["ID","Category","Title","Precondition","Data","Expected","Result"]
RESULT_FILL = {
    "Pass": "C6EFCE", "Fail": "FFC7CE", "Warning": "FFEB9C",
    "Blocked": "D9D9D9", "N/A": "D9D9D9", "Not-run": "D9D9D9",
}

def result_color(v):
    for k, c in RESULT_FILL.items():
        if v.startswith(k):
            return c
    return "FFFFFF"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Login Test Cases"

thin = Side(style="thin", color="BBBBBB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_fill = PatternFill("solid", fgColor="1F4E78")
hdr_font = Font(bold=True, color="FFFFFF")

for c, h in enumerate(HEAD, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.fill = hdr_fill; cell.font = hdr_font; cell.border = border
    cell.alignment = Alignment(horizontal="center", vertical="center")

for r, row in enumerate(CASES, 2):
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    rc = ws.cell(row=r, column=7)
    rc.fill = PatternFill("solid", fgColor=result_color(row[6]))

widths = [10, 15, 34, 18, 22, 40, 22]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:G{len(CASES)+1}"

# Summary sheet
s = wb.create_sheet("Summary")
from collections import Counter
cat = Counter(c[1] for c in CASES)
res = Counter(c[6].split()[0] for c in CASES)
s["A1"] = "By Category"; s["A1"].font = Font(bold=True)
r = 2
for k, v in cat.items():
    s.cell(row=r, column=1, value=k); s.cell(row=r, column=2, value=v); r += 1
r += 1
s.cell(row=r, column=1, value="By Result").font = Font(bold=True); r += 1
for k, v in res.items():
    s.cell(row=r, column=1, value=k); s.cell(row=r, column=2, value=v); r += 1
s.cell(row=r+1, column=1, value="TOTAL"); s.cell(row=r+1, column=2, value=len(CASES))
s.column_dimensions["A"].width = 18

import os
out = os.path.join(os.path.dirname(__file__), "login-testcases.xlsx")
wb.save(out)
print(f"Wrote {out} ({len(CASES)} cases)")
